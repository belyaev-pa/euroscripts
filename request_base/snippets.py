import time
import csv
import openpyxl
import pymorphy2
import nltk
import string
from nltk.corpus import stopwords
from functools import lru_cache, wraps
from .models import *
import pandas as pd
from django.conf import settings



class Namespace: pass
ns = Namespace()
ns.input_list = []
ns.output_list = []
ns.processed_id = []
ns.phrase_to_update = []



def timer(f):
#    @wraps()
    def tmp(*args, **kwargs):
        t = time.time()
        print("start {} at: {}".format(f.__name__, str(time.time() - t)))
        res = f(*args, **kwargs)
        print("ended {} at: {}".format( f.__name__, str(time.time()-t)))
        return res
    return tmp


@timer
def word_tokenize(sentence):
    """take sentence, return list of normalized sentence words"""
    morph = pymorphy2.MorphAnalyzer()
    tokens = nltk.word_tokenize(str(sentence))
    stop_words = stopwords.words('russian')
    stop_words.extend(['что', 'это', 'так', 'вот', 'быть', 'как', 'в', '—', 'к', 'на'])
    tokens = [morph.parse(i)[0].normal_form for i in tokens if (i not in stop_words) and (i not in string.punctuation)]
    return tokens


def get_data(filename1):
    """generator to yield csv row by row"""
    with open(filename1, encoding='Windows-1251') as f_csv:
        reader = csv.reader(f_csv, delimiter=';')
        for row in reader:
            yield (row)


@timer
@lru_cache(maxsize=2)
def get_tag_list(filename):
    """get list of tags and words from xlsx file
    [['tag', list], ['tag', list],...]       """
    wb = openpyxl.load_workbook(filename)
    first_sheet = wb.sheetnames[0]
    worksheet = wb[first_sheet]
    output_list = list()
    for row in range(1, worksheet.max_row+1):
        phrase_cell = "{}{}".format('A', row)
        tag_cell = "{}{}".format('B', row)
        phrase = worksheet[phrase_cell].value
        tag = worksheet[tag_cell].value
        item = list()
        item.append(phrase)
        item.append(tag)
        item.append(word_tokenize(phrase))
        output_list.append(item)
    print('get_tag_list - ended')
    return sorted(output_list, key = lambda s: len(s[1]), reverse=True)


# @timer
# def recursion_sort(phrase, number, level):
#     ns.output_list.append([phrase.phrase, phrase.phrase_len, level+str(number)+'.'])
#     ns.processed_id.append(phrase.id)
#     ns.phrase_to_update.append((phrase.id, level+str(number)))
#     phrase_word_list = phrase.link_set.values_list('word_id', flat=True)
#     phrase_in_phrase = Phrase.objects.filter(
#         phrase_len=phrase.phrase_len+1).exclude(
#         pk__in=ns.processed_id).order_by('phrase_len', '-frequency')
#     for word in phrase_word_list:
#         phrase_in_phrase = phrase_in_phrase.filter(link__word=word)
#     for i, item in enumerate(phrase_in_phrase, 1):
#         recursion_sort(item, i, level+str(number)+'.')
#     return 0


def recursion_sort(obj, number, level):
    s = level+str(number)
    ns.output_list.append((*obj, s))
    ns.input_list.remove(obj)
    ns.phrase_to_update.append((obj[4], s))
    counter = 1
    for item in ns.input_list:
        if obj[1].issubset(item[1]):
            recursion_sort(item, counter, s+'.')
            counter += 1
    return 0


@timer
@lru_cache(maxsize=1024)
def get_phrase_list():
    """
    tuples = [
        (request, set, len, frequency, id),
        (request, set, len, frequency, id),
        (request, set, len, frequency, id),
    ]                                """
    t = time.time()
    print("start - {}".format(str(time.time() - t)))
    tuples_list = list()
    i = 1
    for obj in Phrase.objects.filter(phrase_len__gte=1):
        phrase_word_list = obj.link_set.values_list('word__word', flat=True)
        tuple_element = (obj.phrase, set(phrase_word_list), obj.phrase_len, obj.frequency, obj.id)
        tuples_list.append(tuple_element)
        print('Итерация № {} , я работаю уже: {} с.'.format(i, str(time.time() - t)))
        i += 1
    return sorted(tuples_list, key = lambda x: (x[2], -x[3]))


@timer
def get_phrase_list2():
    l = list(Link.objects.filter(phrase__phrase_len__gte=1).values_list('phrase__phrase',
                                                                        'word__word',
                                                                        'phrase__phrase_len',
                                                                        'phrase__frequency',
                                                                        'phrase__id'))
    ns.df = pd.DataFrame(l, columns=settings.PANDAS_COLUMNS)
    del l
           


def recursion_sort2(row, level, number, sort_str):
    s = sort_str+str(number)+'.'
    print(s)
    ns.output_list.append((row['phrase'].tolist()[0], row['frequency'].tolist()[0], s))
    ns.phrase_to_update.append((row['id'].tolist()[0], s))
    counter = 1
    first_df = ns.df.loc[ns.df['phrase'] == row['phrase'].tolist()[0]]
    word_list = first_df['words'].tolist()
    index_list = first_df.index.tolist()
    ns.df = ns.df.drop(index_list)
    del index_list
    word_df = get_phrase_with_words(word_list, level)
    if word_df.empty:
        word_df = get_phrase_with_words(word_list, level+1)
    del word_list
    phrase_list = word_df.phrase.tolist()
    for phrase in set(phrase_list):
        first = word_df.loc[word_df['phrase'] == phrase].head(1)
        recursion_sort2(first, level+1, counter, s)
        counter += 1
    del phrase_list
    del first_df
    del word_df
    del counter
    return 0


def get_phrase_with_words(word_list, level):
    word_df = pd.DataFrame(columns=settings.PANDAS_COLUMNS)
    append_df = pd.DataFrame(columns=settings.PANDAS_COLUMNS)
    flag = True
    for word in word_list:
        if flag:
            append_df = ns.df.loc[ns.df['words'] == word]
            if not append_df.empty:
                word_df = ns.df.loc[ns.df['phrase'].isin(append_df['phrase'].tolist())]
            flag = False
        else:
            if not word_df.empty:
                append_df = word_df.loc[word_df['words'] == word]
                if not append_df.empty:
                    word_df = ns.df.loc[ns.df['phrase'].isin(append_df['phrase'].tolist())]
    del append_df
    del flag
    return word_df.loc[word_df['phrase_len'] == level+1]
