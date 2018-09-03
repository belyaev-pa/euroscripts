import openpyxl
from django.conf import settings
from django.db.models import Sum
from .models import *
from django.http import HttpResponse
import time


def save_file(mimetype):
    def saver(func):
        def wrapper(request, *args, **kw):
            result = func(request, *args, **kw)
            if isinstance(result, tuple):
                filename, data = result
                resp = HttpResponse(data, content_type=mimetype)
                resp['Content-Disposition'] = 'attachment; filename=%s' % filename
                return resp
            return result
        return wrapper
    return saver


def upload_scheme(file):
    t = time.time()
    print("start time upload_scheme - {}".format(str(time.time() - t)))
    wb = openpyxl.load_workbook(file)
    first_sheet = wb.sheetnames[0]
    worksheet = wb[first_sheet]
    for row in range(2, worksheet.max_row + 1):
        company_id_cell = "{}{}".format('A', row)
        company_cell = "{}{}".format('B', row)
        request_cell = "{}{}".format('C', row)
        url_cell = "{}{}".format('D', row)
        url_type_cell = "{}{}".format('E', row)
        small_change_cell = "{}{}".format('F', row)
        comment_cell = "{}{}".format('G', row)
        head1_cell = "{}{}".format('H', row)
        head2_cell = "{}{}".format('I', row)
        text_cell = "{}{}".format('J', row)
        company_id = worksheet[company_id_cell].value
        company = worksheet[company_cell].value
        request = worksheet[request_cell].value
        url = worksheet[url_cell].value
        url_type = worksheet[url_type_cell].value
        small_change = worksheet[small_change_cell].value
        comment = worksheet[comment_cell].value
        head1 = worksheet[head1_cell].value
        head2 = worksheet[head2_cell].value
        text = worksheet[text_cell].value
        # Если записей будет много необходимо переписать с использованием bulk_create
        url_type_obj, created = PageType.objects.get_or_create(name=url_type)
        url_obj, created = Page.objects.get_or_create(url=url,
                                                      defaults={'page_type': url_type_obj,
                                                                'small_change': small_change,
                                                                'comment': comment, })
        company_obj, created = Company.objects.get_or_create(number=company_id, defaults={'name': company, })
        request_obj, created = Request.objects.get_or_create(name=request, company=company_obj)
        advert_obj, created = Advert.objects.get_or_create(head1=head1, head2=head2,
                                                           defaults={'text': text, })
        ExperimentScheme.objects.get_or_create(request=request_obj,
                                               advert=advert_obj,
                                               page=url_obj, )
    print("end time upload_scheme - {}".format(str(time.time() - t)))


def upload_result(company_file, stat_file):
    t = time.time()
    print("start time upload_result - {}".format(str(time.time() - t)))
    wb = openpyxl.load_workbook(company_file)
    first_sheet = wb.sheetnames[0]
    worksheet = wb[first_sheet]
    company_id_cell = "{}{}".format('E', 8)
    company_id = worksheet[company_id_cell].value
    company_obj = Company.objects.get(number=company_id)
    experiment_scheme = ExperimentScheme.objects.filter(request__company=company_obj)
    request_list = dict()  # {request: list(head1, head2, url)]
    for row in range(12, worksheet.max_row + 1):
        request_cell = "{}{}".format('I', row)
        head1_cell = "{}{}".format('L', row)
        head2_cell = "{}{}".format('M', row)
        url_cell = "{}{}".format('R', row)
        request = worksheet[request_cell].value
        head1 = worksheet[head1_cell].value
        head2 = worksheet[head2_cell].value
        url = worksheet[url_cell].value
        object_request = experiment_scheme.filter(page__url=url,
                                                  advert__head1=head1,
                                                  advert__head2=head2,
                                                  request__name=request)
        if object_request.exists():
            request_list[request] = [head1, head2, url]
    wb = openpyxl.load_workbook(stat_file)
    first_sheet = wb.sheetnames[0]
    worksheet = wb[first_sheet]
    duration_cell = "{}{}".format('A', 3)
    duration = worksheet[duration_cell].value
    for row in range(6, worksheet.max_row + 1):
        request_cell = "{}{}".format('B', row)
        show_cell = "{}{}".format('C', row)
        click_cell = "{}{}".format('D', row)
        ctr_cell = "{}{}".format('E', row)
        expense_cell = "{}{}".format('F', row)
        depth_cell = "{}{}".format('H', row)
        conversion_cell = "{}{}".format('I', row)
        request = worksheet[request_cell].value
        if request not in request_list.keys():
            continue
        show = worksheet[show_cell].value
        click = worksheet[click_cell].value
        ctr = worksheet[ctr_cell].value
        expense = worksheet[expense_cell].value
        depth = worksheet[depth_cell].value
        conversion = worksheet[conversion_cell].value
        experiment_scheme_obj = ExperimentScheme.objects.get(request__name=request,
                                                             advert__head1=request_list[request][0],
                                                             advert__head2=request_list[request][1],
                                                             page__url=request_list[request][2], )
        Experiment.objects.create(scheme=experiment_scheme_obj,
                                  show=show,
                                  click=click,
                                  ctr=ctr,
                                  expense=expense,
                                  depth=depth,
                                  conversion=conversion,
                                  duration=duration, )
    experiment_ended = dict()
    for key, value in request_list.items():
        experiment = ExperimentScheme.objects.get(request__name=key,
                                                  advert__head1=value[0],
                                                  advert__head2=value[1],
                                                  page__url=value[2])
        experiment_list = Experiment.objects.filter(scheme=experiment)
        if experiment_list.aggregate(Sum('show'))['show__sum'] > settings.CLICK_AMOUNT:
            experiment.completed = True
            experiment.save()
            experiment_ended[key] = value
    wb = openpyxl.load_workbook(company_file)
    first_sheet = wb.sheetnames[0]
    worksheet = wb[first_sheet]
    for row in range(12, worksheet.max_row + 1):
        request_cell = "{}{}".format('I', row)
        head1_cell = "{}{}".format('L', row)
        head2_cell = "{}{}".format('M', row)
        text_cell = "{}{}".format('N', row)
        url_cell = "{}{}".format('R', row)
        request = worksheet[request_cell].value
        head1 = worksheet[head1_cell].value
        head2 = worksheet[head2_cell].value
        url = worksheet[url_cell].value
        if request in experiment_ended.keys():
            if len({head1, head2, url} & set(experiment_ended[request])) == 3:
                next_experiment = ExperimentScheme.objects.filter(page__url=experiment_ended[request][2],
                                                                  advert__head1=experiment_ended[request][0],
                                                                  advert__head2=experiment_ended[request][1],
                                                                  request__name=request,
                                                                  completed=False)
                if next_experiment.exists():
                    next_experiment = next_experiment.first()
                    worksheet[head1_cell].value = next_experiment.advert.head1
                    worksheet[head2_cell].value = next_experiment.advert.head2
                    worksheet[text_cell].value = next_experiment.advert.text
                    worksheet[url_cell].value = next_experiment.page.url
    print("end time upload_result - {}".format(str(time.time() - t)))
    return company_file


def make_report():
    wb = openpyxl.load_workbook('report.xlsx')
    first_sheet = wb.sheetnames[0]
    worksheet = wb[first_sheet]
    t = time.time()
    experiment_list = ExperimentScheme.objects.all().order_by('request__company__name')
    worksheet["A1"].value = '№ Компании'
    worksheet["B1"].value = 'Компания'
    worksheet["C1"].value = 'Условие показа'
    worksheet["D1"].value = 'Тип стр'
    worksheet["E1"].value = 'Тип мелкого изменения'
    worksheet["F1"].value = 'Описание мелк изм'
    worksheet["G1"].value = 'URL'
    worksheet["H1"].value = 'заг - 1'
    worksheet["I1"].value = 'заг - 2'
    worksheet["J1"].value = 'объявл'
    worksheet["K1"].value = 'Завершен'
    worksheet["L1"].value = 'Показы_Посл'
    worksheet["M1"].value = 'Клики'
    worksheet["N1"].value = 'CTR (%)'
    worksheet["O1"].value = 'Расход (руб.)'
    worksheet["P1"].value = 'Глубина (стр.)'
    worksheet["Q1"].value = 'Конверсии'
    for row, obj in enumerate(experiment_list, 2):
        company_id_cell = "{}{}".format('A', row)
        company_cell = "{}{}".format('B', row)
        request_cell = "{}{}".format('C', row)
        url_type_cell = "{}{}".format('D', row)
        small_change_cell = "{}{}".format('E', row)
        comment_cell = "{}{}".format('F', row)
        url_cell = "{}{}".format('G', row)
        head1_cell = "{}{}".format('H', row)
        head2_cell = "{}{}".format('I', row)
        text_cell = "{}{}".format('J', row)
        complete_cell = "{}{}".format('K', row)
        show_cell = "{}{}".format('L', row)
        click_cell = "{}{}".format('M', row)
        ctr_cell = "{}{}".format('N', row)
        expense_cell = "{}{}".format('O', row)
        depth_cell = "{}{}".format('P', row)
        conversion_cell = "{}{}".format('Q', row)
        worksheet[company_id_cell].value = obj.request.company.number
        worksheet[company_cell].value = obj.request.company.name
        worksheet[request_cell].value = obj.request.name
        worksheet[url_type_cell].value = obj.page.page_type.name
        worksheet[small_change_cell].value = obj.page.small_change
        worksheet[comment_cell].value = obj.page.comment
        worksheet[url_cell].value = obj.page.url
        worksheet[head1_cell].value = obj.advert.head1
        worksheet[head2_cell].value = obj.advert.head2
        worksheet[text_cell].value = obj.advert.text
        worksheet[complete_cell].value = obj.completed
        experiment_scheme_obj = ExperimentScheme.objects.get(page__url=obj.page.url,
                                                             advert__head1=obj.advert.head1,
                                                             advert__head2=obj.advert.head2,
                                                             request__name=obj.request.name)
        experiment_obj = Experiment.objects.filter(scheme=experiment_scheme_obj)
        worksheet[show_cell].value = experiment_obj.aggregate(Sum('show'))['show__sum']
        worksheet[click_cell].value = experiment_obj.aggregate(Sum('click'))['click__sum']
        worksheet[ctr_cell].value = experiment_obj.aggregate(Sum('ctr'))['ctr__sum']
        worksheet[expense_cell].value = experiment_obj.aggregate(Sum('expense'))['expense__sum']
        worksheet[depth_cell].value = experiment_obj.aggregate(Sum('depth'))['depth__sum']
        worksheet[conversion_cell].value = experiment_obj.aggregate(Sum('conversion'))['conversion__sum']
        print("next {} iter: {}".format(row, str(time.time() - t)))
    wb.save('report.xlsx')
    return open('report.xlsx', 'rb')
